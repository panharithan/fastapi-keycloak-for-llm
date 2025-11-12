import os
import textract
import docx2txt
import openpyxl
from PyPDF2 import PdfReader

def extract_file_content(uploaded_file):
    try:
        # For file-like objects (usual for PDFs, DOCX, XLSX)
        content = uploaded_file.read()
    except AttributeError:
        # For NamedString (string content), just use the string itself
        content = uploaded_file

    # If content is bytes, decode to string (for text files)
    if isinstance(content, bytes):
        try:
            content = content.decode('utf-8')
        except UnicodeDecodeError:
            # If can't decode, just keep as bytes (maybe binary)
            pass
    return content
def extract_text_from_file(file_obj):
    """
    Extract text from a PDF, DOCX, XLSX, or other document using specialized
    parsers with textract fallback. Returns clean string content.
    """
    ext = os.path.splitext(file_obj.name)[1].lower()
    content = ""

    try:
        if ext == ".pdf":
            # Faster and more reliable than textract for PDFs
            reader = PdfReader(file_obj)
            for page in reader.pages:
                content += page.extract_text() or ""

        elif ext == ".docx":
            content = docx2txt.process(file_obj.name)

        elif ext in [".xlsx", ".xls"]:
            wb = openpyxl.load_workbook(file_obj)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    content += " ".join([str(c) for c in row if c]) + "\n"

        else:
            # Try textract for any other file type
            raw_bytes = file_obj.read()
            file_obj.seek(0)
            text = textract.process(file_obj.name, input_encoding='utf-8')
            content = text.decode("utf-8", errors="ignore")

    except Exception as e:
        # Fallback to textract if specialized parser fails
        try:
            text = textract.process(file_obj.name, input_encoding='utf-8')
            content = text.decode("utf-8", errors="ignore")
        except Exception as fallback_error:
            content = f"❌ Error extracting text: {e} / {fallback_error}"

    return content.strip()